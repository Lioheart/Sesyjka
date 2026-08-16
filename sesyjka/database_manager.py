from __future__ import annotations

import logging
import shutil
import sqlite3
import tempfile
import zipfile
from contextlib import closing, contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterator, Sequence

from .config import CORE_DB_FILES, DB_FILES, data_dir


LOG = logging.getLogger(__name__)


class ReadOnlyDatabaseError(PermissionError):
    pass


class DatabaseManager:
    """Zarządza ścieżkami, inicjalizacją i transferem baz Sesyjki."""

    PRIMARY_TABLES = {
        "systemy_rpg.db": "systemy_rpg",
        "sesje_rpg.db": "sesje_rpg",
        "gracze.db": "gracze",
        "wydawcy.db": "wydawcy",
        "planszowe.db": "planszowe",
        "zasoby.db": "zasoby",
    }
    SCHEMA_REQUIREMENTS: dict[str, dict[str, set[str]]] = {
        "wydawcy.db": {
            "wydawcy": {"id", "nazwa", "strona", "kraj"},
        },
        "gracze.db": {
            "gracze": {
                "id",
                "nick",
                "imie_nazwisko",
                "plec",
                "social",
                "glowny_uzytkownik",
                "wazna",
                "grupa",
            },
        },
        "systemy_rpg.db": {
            "systemy_rpg": {
                "id",
                "nazwa",
                "typ",
                "system_glowny_id",
                "typ_suplementu",
                "wydawca_id",
                "fizyczny",
                "pdf",
                "jezyk",
                "status_gra",
                "status_kolekcja",
                "cena_zakupu",
                "waluta_zakupu",
                "cena_sprzedazy",
                "waluta_sprzedazy",
                "vtt",
                "system_glowny_nazwa_custom",
                "system_gry_id",
                "cena_fiz",
                "cena_pdf",
                "cena_vtt",
                "rok_wydania",
                "isbn",
            },
            "systemy_gry": {"id", "nazwa", "wydawca_id", "jezyk", "notatki"},
        },
        "sesje_rpg.db": {
            "sesje_rpg": {
                "id",
                "data_sesji",
                "system_id",
                "liczba_graczy",
                "mg_id",
                "kampania",
                "jednostrzal",
                "tytul_kampanii",
                "tytul_przygody",
                "tryb_gry",
            },
            "sesje_gracze": {"sesja_id", "gracz_id"},
            "sesje_notatki": {"sesja_id", "tresc", "data_modyfikacji"},
        },
        "planszowe.db": {
            "planszowe": {
                "id",
                "nazwa",
                "typ",
                "min_graczy",
                "max_graczy",
                "czas_min",
                "czas_max",
                "minimalny_wiek",
                "cena",
                "waluta",
                "status_gra",
                "status_kolekcja",
                "wydawca_id",
                "wydawca",
                "rok_wydania",
                "notatki",
            },
        },
        "zasoby.db": {
            "zasoby": {
                "id", "pozycja_rpg_id", "typ", "nazwa", "dostawca",
                "format", "sha256", "nazwa_pliku", "tytul_pliku", "external_id",
                "product_url", "rozmiar", "isbn", "wydawca",
                "data_zakupu", "utworzono", "zmodyfikowano",
            },
            "lokalizacje": {
                "id", "zasob_id", "typ", "magazyn_uuid",
                "sciezka_wzgledna", "url", "provider_ref",
                "preferowana", "ostatnio_dostepny",
            },
            "magazyny": {
                "id", "uuid", "nazwa", "typ", "sciezka_bazowa", "aktywny",
            },
        },
    }

    def __init__(self, root: Path | None = None) -> None:
        self._own_root = Path(root) if root is not None else data_dir()
        self._own_root.mkdir(parents=True, exist_ok=True)
        self._guest_root: Path | None = None
        self.last_schema_backup: Path | None = None
        self._write_listeners: list[Callable[[str], None]] = []

    @property
    def own_root(self) -> Path:
        return self._own_root

    @property
    def active_root(self) -> Path:
        return self._guest_root or self._own_root

    @property
    def guest_mode(self) -> bool:
        return self._guest_root is not None

    def enter_guest_mode(self, source: Path) -> None:
        source = Path(source)
        missing = [name for name in CORE_DB_FILES if not (source / name).is_file()]
        if missing:
            raise ValueError("Brak wymaganych baz: " + ", ".join(missing))
        found = [name for name in DB_FILES if (source / name).is_file()]
        self._validate_database_files(source, found, require_current_schema=True)
        self._guest_root = source

    def leave_guest_mode(self) -> None:
        self._guest_root = None

    def path(self, filename: str, own: bool = False) -> Path:
        if filename not in DB_FILES:
            raise ValueError(f"Nieznana baza danych: {filename}")
        return (self._own_root if own else self.active_root) / filename

    def add_write_listener(self, callback: Callable[[str], None]) -> None:
        """Register a listener notified after a committed local database write.

        The listener receives the database filename. It is intentionally outside
        the SQLite transaction, so synchronization bookkeeping can live in
        ``sync.db`` without coupling the domain database schemas to cloud sync.
        """
        if callback not in self._write_listeners:
            self._write_listeners.append(callback)

    def remove_write_listener(self, callback: Callable[[str], None]) -> None:
        try:
            self._write_listeners.remove(callback)
        except ValueError:
            pass

    def _notify_write(self, filename: str) -> None:
        for callback in tuple(self._write_listeners):
            try:
                callback(filename)
            except Exception:
                # The domain transaction is already committed. A bookkeeping
                # failure must not be reported as a failed user save. Cloud sync
                # also compares file fingerprints before the next pass, so the
                # change will be rediscovered even if this notification fails.
                LOG.exception("Nie udało się oznaczyć zmiany bazy %s", filename)

    @contextmanager
    def connect(self, filename: str, write: bool = False) -> Iterator[sqlite3.Connection]:
        if write and self.guest_mode:
            raise ReadOnlyDatabaseError("Tryb gościa jest tylko do odczytu.")
        path = self.path(filename)
        if self.guest_mode and not path.exists():
            raise FileNotFoundError(path)
        if self.guest_mode:
            connection = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
        else:
            connection = sqlite3.connect(path)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        changes_before = connection.total_changes
        committed_change = False
        try:
            yield connection
            if write:
                connection.commit()
                committed_change = connection.total_changes > changes_before
        except Exception:
            if write:
                connection.rollback()
            raise
        finally:
            connection.close()
        if write and committed_change:
            self._notify_write(filename)

    def initialize(self) -> None:
        self.last_schema_backup = self._backup_before_schema_update()
        self._init_publishers()
        self._init_players()
        self._init_systems()
        self._init_sessions()
        self._init_board_games()
        self._init_digital_resources()

    def create_safety_backup(
        self,
        prefix: str,
        filenames: Sequence[str] | None = None,
    ) -> Path:
        """Create a consistent SQLite backup of selected application databases.

        ``sync.db`` may be included even though it is intentionally excluded from
        DB_FILES and user export/import. Safety backups are internal and are used
        to roll back an interrupted cloud synchronization.
        """
        allowed = set(DB_FILES) | {"sync.db"}
        selected = list(filenames or DB_FILES)
        unknown = [name for name in selected if name not in allowed]
        if unknown:
            raise ValueError("Nieznane bazy danych kopii bezpieczeństwa: " + ", ".join(unknown))
        safe_prefix = "".join(ch for ch in str(prefix) if ch.isalnum() or ch in {"-", "_"}).strip("-_") or "backup"
        backup = (
            self._own_root
            / "backups"
            / f"{safe_prefix}-{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}"
        )
        backup.mkdir(parents=True, exist_ok=False)
        try:
            self.extend_safety_backup(backup, selected)
        except Exception:
            shutil.rmtree(backup, ignore_errors=True)
            raise
        return backup

    def extend_safety_backup(self, backup: Path, filenames: Sequence[str]) -> None:
        """Add consistent snapshots to an existing internal backup directory."""
        backup = Path(backup)
        allowed = set(DB_FILES) | {"sync.db"}
        unknown = [name for name in filenames if name not in allowed]
        if unknown:
            raise ValueError("Nieznane bazy danych kopii bezpieczeństwa: " + ", ".join(unknown))
        backup.mkdir(parents=True, exist_ok=True)
        for filename in filenames:
            source = self._own_root / filename
            if not source.is_file():
                continue
            target = backup / filename
            temporary = target.with_name(f".{target.name}.backup-{datetime.now().strftime('%H%M%S%f')}")
            try:
                with closing(sqlite3.connect(f"file:{source}?mode=ro", uri=True)) as source_db:
                    with closing(sqlite3.connect(temporary)) as target_db:
                        source_db.backup(target_db)
                temporary.replace(target)
            finally:
                try:
                    temporary.unlink()
                except FileNotFoundError:
                    pass

    def restore_safety_backup(
        self,
        backup: Path,
        filenames: Sequence[str] | None = None,
    ) -> None:
        """Restore selected databases from an internal safety backup.

        Every source database is checked before any live file is replaced. This
        method must only be called when no write transaction is active.
        """
        backup = Path(backup)
        allowed = set(DB_FILES) | {"sync.db"}
        selected = list(filenames or [path.name for path in backup.glob("*.db")])
        selected = [name for name in selected if name in allowed and (backup / name).is_file()]
        if not selected:
            raise ValueError("Kopia bezpieczeństwa nie zawiera baz do przywrócenia.")

        for filename in selected:
            source = backup / filename
            try:
                with closing(sqlite3.connect(f"file:{source}?mode=ro", uri=True)) as connection:
                    check = connection.execute("PRAGMA quick_check").fetchone()
                    if check is None or str(check[0]).casefold() != "ok":
                        raise ValueError(f"Kopia bazy {filename} nie przeszła kontroli integralności.")
            except sqlite3.DatabaseError as exc:
                raise ValueError(f"Kopia bazy {filename} nie jest poprawną bazą SQLite.") from exc

        staged: dict[str, Path] = {}
        try:
            for filename in selected:
                target = self._own_root / filename
                temporary = target.with_name(f".{target.name}.restore-{datetime.now().strftime('%H%M%S%f')}")
                shutil.copy2(backup / filename, temporary)
                staged[filename] = temporary
            for filename, temporary in staged.items():
                temporary.replace(self._own_root / filename)
        finally:
            for temporary in staged.values():
                try:
                    temporary.unlink()
                except FileNotFoundError:
                    pass

    def prune_safety_backups(self, prefix: str, keep: int = 10) -> None:
        backup_root = self._own_root / "backups"
        if not backup_root.is_dir():
            return
        directories = sorted(
            (path for path in backup_root.iterdir() if path.is_dir() and path.name.startswith(f"{prefix}-")),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        for old in directories[max(0, int(keep)):]:
            shutil.rmtree(old, ignore_errors=True)

    def recover_empty_publishers_from_backup(self) -> Path | None:
        """Recover a clearly accidental empty publisher database.

        Recovery is deliberately conservative. It runs only when the current
        publisher table is empty while other local databases still reference
        publisher identifiers, and only when a backup contains every referenced
        identifier. This avoids resurrecting publishers after an intentional
        cleanup of an otherwise unreferenced database.
        """
        publisher_path = self._own_root / "wydawcy.db"
        if not publisher_path.is_file():
            return None
        try:
            with closing(sqlite3.connect(f"file:{publisher_path}?mode=ro", uri=True)) as connection:
                row = connection.execute("SELECT COUNT(*) FROM wydawcy").fetchone()
                if row is None or int(row[0] or 0) != 0:
                    return None
        except sqlite3.DatabaseError:
            return None

        referenced: set[int] = set()
        queries = (
            ("systemy_rpg.db", "SELECT wydawca_id FROM systemy_rpg WHERE wydawca_id IS NOT NULL"),
            ("systemy_rpg.db", "SELECT wydawca_id FROM systemy_gry WHERE wydawca_id IS NOT NULL"),
            ("planszowe.db", "SELECT wydawca_id FROM planszowe WHERE wydawca_id IS NOT NULL"),
        )
        for filename, query in queries:
            path = self._own_root / filename
            if not path.is_file():
                continue
            try:
                with closing(sqlite3.connect(f"file:{path}?mode=ro", uri=True)) as connection:
                    referenced.update(int(row[0]) for row in connection.execute(query) if row[0] is not None)
            except sqlite3.DatabaseError:
                return None
        if not referenced:
            return None

        backup_root = self._own_root / "backups"
        if not backup_root.is_dir():
            return None
        candidates = sorted(
            (path for path in backup_root.iterdir() if path.is_dir() and (path / "wydawcy.db").is_file()),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        for backup in candidates:
            source = backup / "wydawcy.db"
            try:
                with closing(sqlite3.connect(f"file:{source}?mode=ro", uri=True)) as connection:
                    rows = connection.execute("SELECT id FROM wydawcy").fetchall()
                    ids = {int(row[0]) for row in rows}
                    check = connection.execute("PRAGMA quick_check").fetchone()
                if not referenced.issubset(ids):
                    continue
                if check is None or str(check[0]).casefold() != "ok":
                    continue
            except (sqlite3.DatabaseError, OSError, ValueError, TypeError):
                continue
            self.restore_safety_backup(backup, ("wydawcy.db",))
            return backup
        return None

    def _backup_before_schema_update(self) -> Path | None:
        existing = [self._own_root / name for name in DB_FILES if (self._own_root / name).is_file()]
        if not existing:
            return None
        if not any(self._database_needs_schema_update(path) for path in existing):
            return None
        backup = (
            self._own_root
            / "backups"
            / f"schema-{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}"
        )
        backup.mkdir(parents=True, exist_ok=False)
        for source in existing:
            shutil.copy2(source, backup / source.name)
        return backup

    def _database_needs_schema_update(self, database: Path) -> bool:
        requirements = self.SCHEMA_REQUIREMENTS.get(database.name, {})
        try:
            with closing(sqlite3.connect(f"file:{database}?mode=ro", uri=True)) as connection:
                for table, required_columns in requirements.items():
                    row = connection.execute(
                        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                        (table,),
                    ).fetchone()
                    if row is None:
                        return True
                    existing_columns = self._columns(connection, table)
                    if not required_columns.issubset(existing_columns):
                        return True

                # 0.9.17 wprowadza migrację danych dla rekordów typu Grupa.
                # Grupa ma przechowywać wyłącznie nazwę, typ i system_gry_id.
                # Traktujemy stare dodatkowe wartości jak migrację schematu, aby
                # przed ich usunięciem powstała standardowa kopia bezpieczeństwa.
                if database.name == "systemy_rpg.db":
                    stale_group = connection.execute(
                        """
                        SELECT 1
                        FROM systemy_rpg
                        WHERE LOWER(TRIM(typ)) = 'grupa'
                          AND (
                            system_glowny_id IS NOT NULL
                            OR NULLIF(TRIM(COALESCE(typ_suplementu, '')), '') IS NOT NULL
                            OR wydawca_id IS NOT NULL
                            OR COALESCE(fizyczny, 0) <> 0
                            OR COALESCE(pdf, 0) <> 0
                            OR NULLIF(TRIM(COALESCE(jezyk, '')), '') IS NOT NULL
                            OR NULLIF(TRIM(COALESCE(status_gra, '')), '') IS NOT NULL
                            OR NULLIF(TRIM(COALESCE(status_kolekcja, '')), '') IS NOT NULL
                            OR cena_zakupu IS NOT NULL
                            OR NULLIF(TRIM(COALESCE(waluta_zakupu, '')), '') IS NOT NULL
                            OR cena_sprzedazy IS NOT NULL
                            OR NULLIF(TRIM(COALESCE(waluta_sprzedazy, '')), '') IS NOT NULL
                            OR NULLIF(TRIM(COALESCE(vtt, '')), '') IS NOT NULL
                            OR NULLIF(TRIM(COALESCE(system_glowny_nazwa_custom, '')), '') IS NOT NULL
                            OR cena_fiz IS NOT NULL
                            OR cena_pdf IS NOT NULL
                            OR cena_vtt IS NOT NULL
                            OR rok_wydania IS NOT NULL
                            OR NULLIF(TRIM(COALESCE(isbn, '')), '') IS NOT NULL
                          )
                        LIMIT 1
                        """
                    ).fetchone()
                    if stale_group is not None:
                        return True
        except (sqlite3.DatabaseError, OSError):
            return True
        return False

    @staticmethod
    def _columns(connection: sqlite3.Connection, table: str) -> set[str]:
        return {str(row[1]) for row in connection.execute(f"PRAGMA table_info({table})")}

    @classmethod
    def _ensure_columns(
        cls,
        connection: sqlite3.Connection,
        table: str,
        definitions: dict[str, str],
    ) -> None:
        existing = cls._columns(connection, table)
        for name, sql_type in definitions.items():
            if name not in existing:
                connection.execute(f"ALTER TABLE {table} ADD COLUMN {name} {sql_type}")

    def _init_publishers(self) -> None:
        with self.connect("wydawcy.db", write=True) as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS wydawcy (
                    id INTEGER PRIMARY KEY,
                    nazwa TEXT NOT NULL,
                    strona TEXT,
                    kraj TEXT
                )
                """
            )
            self._ensure_columns(
                connection,
                "wydawcy",
                {"strona": "TEXT", "kraj": "TEXT"},
            )

    def _init_players(self) -> None:
        with self.connect("gracze.db", write=True) as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS gracze (
                    id INTEGER PRIMARY KEY,
                    nick TEXT NOT NULL,
                    imie_nazwisko TEXT,
                    plec TEXT,
                    social TEXT,
                    glowny_uzytkownik INTEGER DEFAULT 0,
                    wazna INTEGER DEFAULT 0,
                    grupa TEXT
                )
                """
            )
            self._ensure_columns(
                connection,
                "gracze",
                {
                    "imie_nazwisko": "TEXT",
                    "plec": "TEXT",
                    "social": "TEXT",
                    "glowny_uzytkownik": "INTEGER DEFAULT 0",
                    "wazna": "INTEGER DEFAULT 0",
                    "grupa": "TEXT",
                },
            )

    def _init_systems(self) -> None:
        with self.connect("systemy_rpg.db", write=True) as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS systemy_rpg (
                    id INTEGER PRIMARY KEY,
                    nazwa TEXT NOT NULL,
                    typ TEXT NOT NULL,
                    system_glowny_id INTEGER,
                    typ_suplementu TEXT,
                    wydawca_id INTEGER,
                    fizyczny INTEGER DEFAULT 0,
                    pdf INTEGER DEFAULT 0,
                    jezyk TEXT,
                    status_gra TEXT DEFAULT 'Nie grane',
                    status_kolekcja TEXT DEFAULT 'W kolekcji',
                    cena_zakupu REAL,
                    waluta_zakupu TEXT,
                    cena_sprzedazy REAL,
                    waluta_sprzedazy TEXT,
                    vtt TEXT,
                    system_glowny_nazwa_custom TEXT,
                    system_gry_id INTEGER,
                    cena_fiz REAL,
                    cena_pdf REAL,
                    cena_vtt REAL,
                    rok_wydania INTEGER,
                    isbn TEXT,
                    FOREIGN KEY (system_glowny_id) REFERENCES systemy_rpg(id)
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS systemy_gry (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nazwa TEXT NOT NULL,
                    wydawca_id INTEGER,
                    jezyk TEXT,
                    notatki TEXT
                )
                """
            )
            self._ensure_columns(
                connection,
                "systemy_rpg",
                {
                    "system_glowny_id": "INTEGER",
                    "typ_suplementu": "TEXT",
                    "wydawca_id": "INTEGER",
                    "fizyczny": "INTEGER DEFAULT 0",
                    "pdf": "INTEGER DEFAULT 0",
                    "jezyk": "TEXT",
                    "status_gra": "TEXT DEFAULT 'Nie grane'",
                    "status_kolekcja": "TEXT DEFAULT 'W kolekcji'",
                    "cena_zakupu": "REAL",
                    "waluta_zakupu": "TEXT",
                    "cena_sprzedazy": "REAL",
                    "waluta_sprzedazy": "TEXT",
                    "vtt": "TEXT",
                    "system_glowny_nazwa_custom": "TEXT",
                    "system_gry_id": "INTEGER",
                    "cena_fiz": "REAL",
                    "cena_pdf": "REAL",
                    "cena_vtt": "REAL",
                    "rok_wydania": "INTEGER",
                    "isbn": "TEXT",
                },
            )
            connection.execute(
                """
                UPDATE systemy_rpg
                SET system_glowny_id = NULL,
                    typ_suplementu = NULL,
                    wydawca_id = NULL,
                    fizyczny = 0,
                    pdf = 0,
                    jezyk = NULL,
                    status_gra = NULL,
                    status_kolekcja = NULL,
                    cena_zakupu = NULL,
                    waluta_zakupu = NULL,
                    cena_sprzedazy = NULL,
                    waluta_sprzedazy = NULL,
                    vtt = NULL,
                    system_glowny_nazwa_custom = NULL,
                    cena_fiz = NULL,
                    cena_pdf = NULL,
                    cena_vtt = NULL,
                    rok_wydania = NULL,
                    isbn = NULL
                WHERE LOWER(TRIM(typ)) = 'grupa'
                  AND (
                    system_glowny_id IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(typ_suplementu, '')), '') IS NOT NULL
                    OR wydawca_id IS NOT NULL
                    OR COALESCE(fizyczny, 0) <> 0
                    OR COALESCE(pdf, 0) <> 0
                    OR NULLIF(TRIM(COALESCE(jezyk, '')), '') IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(status_gra, '')), '') IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(status_kolekcja, '')), '') IS NOT NULL
                    OR cena_zakupu IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(waluta_zakupu, '')), '') IS NOT NULL
                    OR cena_sprzedazy IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(waluta_sprzedazy, '')), '') IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(vtt, '')), '') IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(system_glowny_nazwa_custom, '')), '') IS NOT NULL
                    OR cena_fiz IS NOT NULL
                    OR cena_pdf IS NOT NULL
                    OR cena_vtt IS NOT NULL
                    OR rok_wydania IS NOT NULL
                    OR NULLIF(TRIM(COALESCE(isbn, '')), '') IS NOT NULL
                  )
                """
            )

    def _init_sessions(self) -> None:
        with self.connect("sesje_rpg.db", write=True) as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS sesje_rpg (
                    id INTEGER PRIMARY KEY,
                    data_sesji TEXT NOT NULL,
                    system_id INTEGER NOT NULL,
                    liczba_graczy INTEGER NOT NULL,
                    mg_id INTEGER,
                    kampania INTEGER DEFAULT 0,
                    jednostrzal INTEGER DEFAULT 0,
                    tytul_kampanii TEXT,
                    tytul_przygody TEXT,
                    tryb_gry TEXT
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS sesje_gracze (
                    sesja_id INTEGER NOT NULL,
                    gracz_id INTEGER NOT NULL,
                    PRIMARY KEY (sesja_id, gracz_id),
                    FOREIGN KEY (sesja_id) REFERENCES sesje_rpg(id) ON DELETE CASCADE
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS sesje_notatki (
                    sesja_id INTEGER PRIMARY KEY,
                    tresc TEXT NOT NULL,
                    data_modyfikacji TEXT NOT NULL,
                    FOREIGN KEY (sesja_id) REFERENCES sesje_rpg(id) ON DELETE CASCADE
                )
                """
            )
            self._ensure_columns(
                connection,
                "sesje_rpg",
                {
                    "mg_id": "INTEGER",
                    "kampania": "INTEGER DEFAULT 0",
                    "jednostrzal": "INTEGER DEFAULT 0",
                    "tytul_kampanii": "TEXT",
                    "tytul_przygody": "TEXT",
                    "tryb_gry": "TEXT",
                },
            )


    def _init_board_games(self) -> None:
        publisher_ids = {
            str(row["nazwa"]).strip().casefold(): int(row["id"])
            for row in self.table_rows(
                "wydawcy.db",
                "SELECT id, nazwa FROM wydawcy",
            )
            if str(row["nazwa"] or "").strip()
        }
        with self.connect("planszowe.db", write=True) as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS planszowe (
                    id INTEGER PRIMARY KEY,
                    nazwa TEXT NOT NULL,
                    typ TEXT NOT NULL DEFAULT 'Gra planszowa',
                    min_graczy INTEGER NOT NULL DEFAULT 1,
                    max_graczy INTEGER NOT NULL DEFAULT 1,
                    czas_min INTEGER,
                    czas_max INTEGER,
                    minimalny_wiek INTEGER,
                    cena REAL,
                    waluta TEXT DEFAULT 'PLN',
                    status_gra TEXT DEFAULT 'Nie grane',
                    status_kolekcja TEXT DEFAULT 'W kolekcji',
                    wydawca_id INTEGER,
                    wydawca TEXT,
                    rok_wydania INTEGER,
                    notatki TEXT
                )
                """
            )
            self._ensure_columns(
                connection,
                "planszowe",
                {
                    "typ": "TEXT NOT NULL DEFAULT 'Gra planszowa'",
                    "min_graczy": "INTEGER NOT NULL DEFAULT 1",
                    "max_graczy": "INTEGER NOT NULL DEFAULT 1",
                    "czas_min": "INTEGER",
                    "czas_max": "INTEGER",
                    "minimalny_wiek": "INTEGER",
                    "cena": "REAL",
                    "waluta": "TEXT DEFAULT 'PLN'",
                    "status_gra": "TEXT DEFAULT 'Nie grane'",
                    "status_kolekcja": "TEXT DEFAULT 'W kolekcji'",
                    "wydawca_id": "INTEGER",
                    "wydawca": "TEXT",
                    "rok_wydania": "INTEGER",
                    "notatki": "TEXT",
                },
            )

            # Wersje 0.8.0-0.8.3 przechowywały wydawcę wyłącznie jako tekst.
            # Relacja między osobnymi plikami SQLite jest egzekwowana w Pythonie.
            # Przy migracji wiążemy dokładnie pasujące nazwy bez usuwania pola
            # tekstowego, aby starsze wydania nadal mogły odczytać bazę.
            legacy_rows = connection.execute(
                """
                SELECT id, wydawca
                FROM planszowe
                WHERE wydawca_id IS NULL AND TRIM(COALESCE(wydawca, '')) <> ''
                """
            ).fetchall()
            for row in legacy_rows:
                publisher_id = publisher_ids.get(str(row["wydawca"]).strip().casefold())
                if publisher_id is not None:
                    connection.execute(
                        "UPDATE planszowe SET wydawca_id=? WHERE id=?",
                        (publisher_id, int(row["id"])),
                    )


    def _init_digital_resources(self) -> None:
        with self.connect("zasoby.db", write=True) as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS zasoby (
                    id INTEGER PRIMARY KEY,
                    pozycja_rpg_id INTEGER,
                    typ TEXT NOT NULL DEFAULT 'PDF',
                    nazwa TEXT NOT NULL,
                    dostawca TEXT,
                    format TEXT,
                    sha256 TEXT,
                    nazwa_pliku TEXT,
                    tytul_pliku TEXT,
                    external_id TEXT,
                    product_url TEXT,
                    rozmiar INTEGER,
                    isbn TEXT,
                    wydawca TEXT,
                    data_zakupu TEXT,
                    utworzono TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    zmodyfikowano TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );

                CREATE UNIQUE INDEX IF NOT EXISTS idx_zasoby_external_id
                ON zasoby(external_id) WHERE external_id IS NOT NULL;

                CREATE INDEX IF NOT EXISTS idx_zasoby_pozycja_rpg
                ON zasoby(pozycja_rpg_id);

                CREATE TABLE IF NOT EXISTS lokalizacje (
                    id INTEGER PRIMARY KEY,
                    zasob_id INTEGER NOT NULL,
                    typ TEXT NOT NULL,
                    magazyn_uuid TEXT,
                    sciezka_wzgledna TEXT,
                    url TEXT,
                    provider_ref TEXT,
                    preferowana INTEGER NOT NULL DEFAULT 0,
                    ostatnio_dostepny INTEGER NOT NULL DEFAULT 0,
                    FOREIGN KEY (zasob_id) REFERENCES zasoby(id) ON DELETE CASCADE
                );

                CREATE INDEX IF NOT EXISTS idx_lokalizacje_zasob
                ON lokalizacje(zasob_id);

                CREATE TABLE IF NOT EXISTS magazyny (
                    id INTEGER PRIMARY KEY,
                    uuid TEXT NOT NULL UNIQUE,
                    nazwa TEXT NOT NULL,
                    typ TEXT NOT NULL DEFAULT 'Lokalny',
                    sciezka_bazowa TEXT NOT NULL,
                    aktywny INTEGER NOT NULL DEFAULT 1
                );
                """
            )
            self._ensure_columns(
                connection,
                "zasoby",
                {
                    "pozycja_rpg_id": "INTEGER",
                    "typ": "TEXT NOT NULL DEFAULT 'PDF'",
                    "nazwa": "TEXT",
                    "dostawca": "TEXT",
                    "format": "TEXT",
                    "sha256": "TEXT",
                    "nazwa_pliku": "TEXT",
                    "tytul_pliku": "TEXT",
                    "external_id": "TEXT",
                    "product_url": "TEXT",
                    "rozmiar": "INTEGER",
                    "isbn": "TEXT",
                    "wydawca": "TEXT",
                    "data_zakupu": "TEXT",
                    "utworzono": "TEXT",
                    "zmodyfikowano": "TEXT",
                },
            )
            self._ensure_columns(
                connection,
                "lokalizacje",
                {
                    "magazyn_uuid": "TEXT",
                    "sciezka_wzgledna": "TEXT",
                    "url": "TEXT",
                    "provider_ref": "TEXT",
                    "preferowana": "INTEGER NOT NULL DEFAULT 0",
                    "ostatnio_dostepny": "INTEGER NOT NULL DEFAULT 0",
                },
            )
            self._ensure_columns(
                connection,
                "magazyny",
                {
                    "uuid": "TEXT",
                    "nazwa": "TEXT",
                    "typ": "TEXT NOT NULL DEFAULT 'Lokalny'",
                    "sciezka_bazowa": "TEXT",
                    "aktywny": "INTEGER NOT NULL DEFAULT 1",
                },
            )

    def has_active_database(self, filename: str) -> bool:
        if filename not in DB_FILES:
            return False
        return self.path(filename).is_file()

    def next_id(self, filename: str, table: str) -> int:
        with self.connect(filename) as connection:
            row = connection.execute(f"SELECT MAX(id) AS max_id FROM {table}").fetchone()
            return int(row["max_id"] or 0) + 1

    def export_zip(self, destination: Path) -> Path:
        destination = Path(destination)
        if destination.suffix.lower() != ".zip":
            destination = destination.with_suffix(".zip")
        destination.parent.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(destination, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for filename in DB_FILES:
                source = self.path(filename, own=True)
                if source.exists():
                    archive.write(source, arcname=filename)
        return destination

    def export_folder(self, destination: Path) -> Path:
        """Copy all application databases into a selected directory."""
        destination = Path(destination)
        destination.mkdir(parents=True, exist_ok=True)
        for filename in DB_FILES:
            source = self.path(filename, own=True)
            if source.exists():
                shutil.copy2(source, destination / filename)
        return destination

    def export_excel(self, destination: Path) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("Eksport XLSX wymaga pakietu openpyxl.") from exc

        destination = Path(destination)
        if destination.suffix.lower() != ".xlsx":
            destination = destination.with_suffix(".xlsx")
        workbook = Workbook()
        workbook.remove(workbook.active)
        labels = {
            "systemy_rpg.db": "Systemy RPG",
            "sesje_rpg.db": "Sesje RPG",
            "gracze.db": "Gracze",
            "wydawcy.db": "Wydawcy",
            "planszowe.db": "Gry planszowe",
            "zasoby.db": "Zasoby cyfrowe",
        }
        for filename in DB_FILES:
            source = self.path(filename, own=True)
            if not source.exists():
                continue
            with closing(sqlite3.connect(source)) as connection:
                connection.row_factory = sqlite3.Row
                tables = connection.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
                ).fetchall()
                for table_row in tables:
                    table = str(table_row["name"])
                    title = f"{labels[filename]} - {table}"[:31]
                    sheet = workbook.create_sheet(title)
                    rows = connection.execute(f'SELECT * FROM "{table}"').fetchall()
                    headers = [row[1] for row in connection.execute(f'PRAGMA table_info("{table}")')]
                    sheet.append(headers)
                    for cell in sheet[1]:
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.fill = PatternFill("solid", fgColor="1565C0")
                        cell.alignment = Alignment(horizontal="center")
                    for row in rows:
                        sheet.append([row[header] for header in headers])
                    for column in sheet.columns:
                        width = min(50, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
                        sheet.column_dimensions[column[0].column_letter].width = width
        if not workbook.sheetnames:
            workbook.create_sheet("Brak danych")
        workbook.save(destination)
        return destination

    def inspect_import_source(self, source: Path) -> tuple[Path, list[str], Path | None]:
        source = Path(source)
        cleanup: Path | None = None
        if source.is_dir():
            root = source
        elif source.suffix.lower() == ".zip":
            cleanup = Path(tempfile.mkdtemp(prefix="sesyjka-import-"))
            with zipfile.ZipFile(source) as archive:
                names = {Path(name).name for name in archive.namelist()}
                allowed = [name for name in DB_FILES if name in names]
                if not allowed:
                    shutil.rmtree(cleanup, ignore_errors=True)
                    raise ValueError("Archiwum nie zawiera baz danych Sesyjki.")
                for name in allowed:
                    member = next(item for item in archive.namelist() if Path(item).name == name)
                    with archive.open(member) as src, (cleanup / name).open("wb") as dst:
                        shutil.copyfileobj(src, dst)
            root = cleanup
        else:
            raise ValueError("Wybierz katalog albo archiwum ZIP.")
        found = [name for name in DB_FILES if (root / name).is_file()]
        if not found:
            if cleanup:
                shutil.rmtree(cleanup, ignore_errors=True)
            raise ValueError("Nie znaleziono baz danych Sesyjki.")
        try:
            self._validate_database_files(root, found)
        except ValueError:
            if cleanup:
                shutil.rmtree(cleanup, ignore_errors=True)
            raise
        return root, found, cleanup

    def _validate_database_files(
        self,
        root: Path,
        filenames: Sequence[str],
        *,
        require_current_schema: bool = False,
    ) -> None:
        try:
            for filename in filenames:
                database = Path(root) / filename
                tables = (
                    self.SCHEMA_REQUIREMENTS[filename]
                    if require_current_schema
                    else {self.PRIMARY_TABLES[filename]: set()}
                )
                with closing(sqlite3.connect(f"file:{database}?mode=ro", uri=True)) as connection:
                    for table, required_columns in tables.items():
                        row = connection.execute(
                            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                            (table,),
                        ).fetchone()
                        if row is None:
                            raise ValueError(
                                f"Plik {filename} nie zawiera wymaganej tabeli {table}."
                            )
                        if required_columns:
                            existing_columns = self._columns(connection, table)
                            if not required_columns.issubset(existing_columns):
                                raise ValueError(
                                    f"Plik {filename} ma niezgodny schemat tabeli {table}."
                                )
                    check = connection.execute("PRAGMA quick_check").fetchone()
                    if check is None or str(check[0]).casefold() != "ok":
                        raise ValueError(
                            f"Kontrola integralności bazy {filename} nie zakończyła się poprawnie."
                        )
        except ValueError:
            raise
        except (sqlite3.DatabaseError, OSError) as exc:
            raise ValueError(
                "Co najmniej jeden plik nie jest poprawną bazą SQLite Sesyjki."
            ) from exc

    def replace_own_databases(self, source: Path, filenames: Sequence[str]) -> Path:
        if self.guest_mode:
            raise ReadOnlyDatabaseError("Najpierw zakończ tryb gościa.")
        backup = self._own_root / "backups" / datetime.now().strftime("%Y%m%d-%H%M%S")
        backup.mkdir(parents=True, exist_ok=True)
        for filename in filenames:
            if filename not in DB_FILES:
                continue
            current = self.path(filename, own=True)
            incoming = Path(source) / filename
            if current.exists():
                shutil.copy2(current, backup / filename)
            shutil.copy2(incoming, current)
        self.initialize()
        return backup

    def table_rows(self, filename: str, query: str, parameters: Sequence[Any] = ()) -> list[sqlite3.Row]:
        with self.connect(filename) as connection:
            return list(connection.execute(query, parameters).fetchall())
